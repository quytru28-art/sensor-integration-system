#!/usr/bin/env python3
"""Align UserStories_2026_03_11.xlsx with the implemented sensor-integration-system."""
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "UserStories_2026_03_11.xlsx"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)

    us = wb["User Stories"]
    # (excel_row, col_G_status, col_E_ac_optional) — row numbers from sheet inspection
    updates = {
        6: {  # ID 3
            "status": "Complete",
            "ac": "User can set a display name and device type for each sensor; both are shown on device cards (no separate long-text description field).",
        },
        7: {  # ID 4
            "status": "Complete",
            "ac": "User can set a new password from Profile (with confirmation); complexity rules enforced; re-login requires new password. (API: new password only — no current-password field.)",
        },
        16: {  # ID 13
            "status": "Complete",
            "ac": "User receives an email alert on failed password attempts (optional toggle in Profile). Admin inbox receives forgot-password notifications.",
        },
        17: {  # ID 14
            "status": "Incomplete",
            "ac": "Not implemented: SMS/email-delivered MFA. The app uses an in-browser 6-character verification step on login/signup only.",
        },
        19: {  # ID 16
            "status": "Complete",
            "ac": "User can view and edit collection interval (minutes) per device in Edit Device; persisted via PATCH /api/devices/:id.",
        },
        22: {  # ID 19
            "status": "Complete",
            "ac": "User can view line charts (temperature, humidity) for historical readings when View Data is open; Chart.js updates with filtered data.",
        },
        24: {  # ID 21
            "status": "Complete",
            "ac": "User can update email and phone from Profile; saved via PATCH /api/account.",
        },
        25: {  # ID 22
            "status": "Complete",
            "ac": "Email alerts sent when sensor readings breach configured thresholds (and on failed logins if enabled). SMS not implemented.",
        },
    }

    for row, spec in updates.items():
        us.cell(row, 7, spec["status"])
        if "ac" in spec:
            us.cell(row, 5, spec["ac"])

    # Fix duplicate ID: Self-deactivate row had ID 32; QR scan is the real #32
    us.cell(38, 1, 39)  # column A = ID

    su = wb["Summary"]
    total = 39
    complete = 31
    incomplete = 8
    pct_c = round(complete / total * 100)
    pct_i = round(incomplete / total * 100)
    su["B3"] = total
    su["B4"] = complete
    su["C4"] = f"{pct_c}%"
    su["B6"] = incomplete
    su["C6"] = f"{pct_i}%"
    # Must / Should / Could complete counts
    su["B9"], su["C9"], su["D9"] = 17, 17, "100%"
    su["B10"], su["C10"], su["D10"] = 15, 9, "60%"
    su["B11"], su["C11"], su["D11"] = 7, 5, "71%"
    su["A13"] = (
        f"✨  {complete} of {total} stories Complete  ·  Sprint 3 of 3 done  ·  "
        f"{incomplete} PBI remaining  ·  Avg velocity: {complete/3:.1f} PBI/sprint"
    )

    sr = wb["Sprint Review"]
    sr["D3"] = total
    sr["D4"] = complete
    sr["D5"] = incomplete
    sr["D6"] = round(complete / 3, 1)
    # Sprint 3 summary row 18
    sr["D18"] = complete
    sr["E18"] = incomplete
    sr["F18"] = (
        "Sensor log, date filter, export CSV/JSON, QR scan, thresholds, threshold & failed-login email alerts, "
        "connectivity status, data retention, edit sensor, collection interval, charts, profile email/phone"
    )
    # Sprint 1 row still lists #14 as delivered — clarify scope
    sr["B31"] = (
        "#14 Verification step on login/signup (not SMS/email MFA — full MFA remains Incomplete in User Stories sheet)"
    )

    wb.save(XLSX)
    print(f"Updated {XLSX}")


if __name__ == "__main__":
    main()
